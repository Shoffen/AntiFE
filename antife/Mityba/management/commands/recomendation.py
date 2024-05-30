from django.core.management.base import BaseCommand
import openpyxl
from collections import Counter

class Command(BaseCommand):
    help = 'Export data to Excel'

    def handle(self, *args, **kwargs):
        # Export logic
        def export_to_excel():
            # Load the original workbook
            wb = openpyxl.load_workbook('valgiarastis_data.xlsx')

            # Create a new workbook for rekomendacijos
            rekomendacijos_wb = openpyxl.Workbook()

            # Iterate through each sheet in the original workbook
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Create a new sheet in the rekomendacijos workbook
                rekomendacijos_ws = rekomendacijos_wb.create_sheet(title=sheet_name)

                # Initialize a Counter object to count the occurrences of Pavadinimas
                counter = Counter()

                # Iterate through each row in the sheet, skipping the header row
                for row in ws.iter_rows(min_row=2, values_only=True):
                    # Skip empty rows
                    if any(row):
                        # Count the occurrences of Pavadinimas
                        pavadinimas_cell = row[0]  # Assuming Pavadinimas is the first column
                        if isinstance(pavadinimas_cell, str):
                            counter.update([pavadinimas_cell])

                # Write the headers
                rekomendacijos_ws.append(["Pavadinimas", "Kiekis"])

                # Get the three most common elements
                most_common = counter.most_common(3)

                # Write the most common elements into the new sheet
                for (pavadinimas, count) in most_common:
                    rekomendacijos_ws.append([pavadinimas, count])

            # Remove the default empty sheet
            rekomendacijos_wb.remove(rekomendacijos_wb.active)

            # Save the new workbook
            rekomendacijos_wb.save('rekomendacijos.xlsx')

        # Call export_to_excel function
        export_to_excel()
        self.stdout.write(self.style.SUCCESS('Data exported successfully!'))
