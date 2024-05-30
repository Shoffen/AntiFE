from django.core.management.base import BaseCommand
import openpyxl
from homepage.models import Valgomas_produktas, Valgymo_receptas, Receptai, Product, Valgymai

class Command(BaseCommand):
    help = 'Export data to Excel'

    def handle(self, *args, **kwargs):
        # Export logic
        def export_to_excel():
            # Retrieve data from the database
            valgomas_produktas_data = Valgomas_produktas.objects.all()
            valgymo_receptas_data = Valgymo_receptas.objects.all()

            # Create a new Excel workbook
            wb = openpyxl.Workbook()

            # Initialize variables to track if any data exists for each meal type
            pusryciai_exists = False
            pietus_exists = False
            vakariene_exists = False
            papildomai_exists = False

            # Write data for Valgomas_produktas to appropriate sheets
            for item in valgomas_produktas_data:
                valgymas = Valgymai.objects.get(pk=item.fk_Valgymasid_Valgymas_id)
                if valgymas.tipas == "Pusryčiai":
                    pusryciai_exists = True
                elif valgymas.tipas == "Pietūs":
                    pietus_exists = True
                elif valgymas.tipas == "Vakarienė":
                    vakariene_exists = True
                else:
                    papildomai_exists = True

            # Write data for Valgymo_receptas to appropriate sheets
            for item in valgymo_receptas_data:
                valgymas = Valgymai.objects.get(pk=item.fk_Valgymasid_Valgymas_id)
                if valgymas.tipas == "Pusryčiai":
                    pusryciai_exists = True
                elif valgymas.tipas == "Pietūs":
                    pietus_exists = True
                elif valgymas.tipas == "Vakarienė":
                    vakariene_exists = True
                else:
                    papildomai_exists = True

            # Create sheets for meal types with data
            if pusryciai_exists:
                ws_pusryciai = wb.create_sheet("Pusryčiai")
                ws_pusryciai.append(["Pavadinimas", "Kategorija", "Kiekis"])
            if pietus_exists:
                ws_pietus = wb.create_sheet("Pietūs")
                ws_pietus.append(["Pavadinimas", "Kategorija", "Kiekis"])
            if vakariene_exists:
                ws_vakariene = wb.create_sheet("Vakarienė")
                ws_vakariene.append(["Pavadinimas", "Kategorija", "Kiekis"])
            if papildomai_exists:
                ws_papildomai = wb.create_sheet("Papildomai")
                ws_papildomai.append(["Pavadinimas", "Kategorija", "Kiekis"])

            # Write data to appropriate sheets
            for item in valgomas_produktas_data:
                produktas = Product.objects.get(pk=item.fk_Produktasid_Produktas_id)
                valgymas = Valgymai.objects.get(pk=item.fk_Valgymasid_Valgymas_id)
                if valgymas.tipas == "Pusryčiai":
                    ws_pusryciai.append([produktas.name, valgymas.tipas, item.kiekis])
                elif valgymas.tipas == "Pietūs":
                    ws_pietus.append([produktas.name, valgymas.tipas, item.kiekis])
                elif valgymas.tipas == "Vakarienė":
                    ws_vakariene.append([produktas.name, valgymas.tipas, item.kiekis])
                else:
                    ws_papildomai.append([produktas.name, valgymas.tipas, item.kiekis])

            for item in valgymo_receptas_data:
                receptas = Receptai.objects.get(pk=item.fk_Receptasid_Receptas_id)
                valgymas = Valgymai.objects.get(pk=item.fk_Valgymasid_Valgymas_id)
                if valgymas.tipas == "Pusryčiai":
                    ws_pusryciai.append([receptas.pavadinimas, valgymas.tipas, item.kiekis])
                elif valgymas.tipas == "Pietūs":
                    ws_pietus.append([receptas.pavadinimas, valgymas.tipas, item.kiekis])
                elif valgymas.tipas == "Vakarienė":
                    ws_vakariene.append([receptas.pavadinimas, valgymas.tipas, item.kiekis])
                else:
                    ws_papildomai.append([receptas.pavadinimas, valgymas.tipas, item.kiekis])

            # Save the workbook
            wb.remove(wb.worksheets[0])
            wb.save("valgiarastis_data.xlsx")

        # Call export_to_excel function
        export_to_excel()
        self.stdout.write(self.style.SUCCESS('Data exported successfully!'))

