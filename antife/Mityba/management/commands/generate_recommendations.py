# <your_app>/management/commands/generate_recommendations.py

import os
import random
import openpyxl
import pandas as pd
from django.conf import settings
from django.core.management.base import BaseCommand
from homepage.models import Receptai, Product
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class Command(BaseCommand):
    help = 'Generates Excel file for recommendations'

    def add_arguments(self, parser):
        parser.add_argument('bendras_baltymas', type=str, help='Total protein amount')
        parser.add_argument('bendras_fenilalaninas', type=str, help='Total phenylalanine amount')

    def generate_recommendations(self, bendras_baltymas, bendras_fenilalaninas):
        print(f"Bendras baltymu kiekis: {bendras_baltymas}")
        print(f"Bendras fenilalanino kiekis: {bendras_fenilalaninas}")

        categories = ['Pusryčiai', 'Pietus', 'Vakarienė', 'Papildomai']
        recommendations = {}

        recommendation = random.choice(Receptai.objects.filter(visible=True))
        recommendations['Pusryčiai'] = f"Receptas: {recommendation.pavadinimas}"

        for category in categories[1:]:
            if random.choice([True, False]):
                recommendation = random.choice(Receptai.objects.filter(visible=True))
                recommendations[category.lower()] = f"Receptas: {recommendation.pavadinimas}"
            else:
                recommendation = random.choice(Product.objects.all())
                recommendations[category.lower()] = f"Produktas: {recommendation.name}"

        df = pd.DataFrame(recommendations.items(), columns=['Category', 'Recommendation'])

        file_path = os.path.join(settings.BASE_DIR, 'rekomendacijos.xlsx')
        wb = load_workbook(filename=file_path)

        try:
            ws = wb['Individualios_rek']
        except KeyError:
            ws = wb.create_sheet("Individualios_rek")

        ws.cell(row=1, column=1, value="").fill = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for row_num, (_, row) in enumerate(df.iterrows(), start=2):
            for col_num, value in enumerate(row):
                ws.cell(row=row_num, column=col_num + 1, value=value)

        header_font = Font(bold=True)
        for col_num, _ in enumerate(df.columns):
            ws.cell(row=1, column=col_num + 1, value=df.columns[col_num]).font = header_font

        wb.save(filename=file_path)
        return file_path

    def handle(self, *args, **options):
        bendras_baltymas = options['bendras_baltymas']
        bendras_fenilalaninas = options['bendras_fenilalaninas']
        self.generate_recommendations(bendras_baltymas, bendras_fenilalaninas)
        self.stdout.write(self.style.SUCCESS('Successfully generated recommendations Excel file'))
